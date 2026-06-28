"""Task analytics aggregation + Excel export."""
from __future__ import annotations

import io
import unittest
from contextlib import contextmanager
from datetime import datetime, timedelta

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models import CommunicationTask  # noqa: F401
from app.services import task_analytics_service as analytics


@contextmanager
def _temp_db():
    engine = create_engine("sqlite:///:memory:", future=True)
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    db = Session()
    try:
        yield db
    finally:
        db.close()
        engine.dispose()


def _seed(db):
    past = datetime.utcnow() - timedelta(days=2)
    db.add_all([
        CommunicationTask(title="a", status="TODO", priority="P1", task_source="SUPPLIER",
                          assigned_to_user_id=1, assigned_to="Alice", due_date=past),
        CommunicationTask(title="b", status="DONE", priority="P2", task_source="CUSTOMER",
                          assigned_to_user_id=1, assigned_to="Alice", closed_at=datetime.utcnow()),
        CommunicationTask(title="c", status="IN_PROGRESS", priority="P0", task_source="INTERNAL",
                          assigned_to_user_id=2, assigned_to="Bob"),
    ])
    db.commit()


class AnalyticsTests(unittest.TestCase):
    def test_totals_and_breakdowns(self) -> None:
        with _temp_db() as db:
            _seed(db)
            data = analytics.compute_analytics(db)
            self.assertEqual(data["totals"]["total"], 3)
            self.assertEqual(data["totals"]["done"], 1)
            self.assertEqual(data["totals"]["open"], 2)
            self.assertEqual(data["totals"]["overdue"], 1)
            self.assertEqual(data["by_status"]["TODO"], 1)
            self.assertEqual(data["by_priority"]["P0"], 1)
            self.assertEqual(data["by_source"]["CUSTOMER"], 1)

    def test_by_assignee_groups_real_users(self) -> None:
        with _temp_db() as db:
            _seed(db)
            data = analytics.compute_analytics(db)
            by = {r["name"]: r for r in data["by_assignee"]}
            self.assertEqual(by["Alice"]["open"], 1)
            self.assertEqual(by["Alice"]["done"], 1)
            self.assertEqual(by["Bob"]["open"], 1)

    def test_export_workbook_is_valid_xlsx(self) -> None:
        from openpyxl import load_workbook
        with _temp_db() as db:
            _seed(db)
            data = analytics.export_workbook(db)
            wb = load_workbook(io.BytesIO(data))
            self.assertIn("Tasks", wb.sheetnames)
            self.assertIn("Activity", wb.sheetnames)
            ws = wb["Tasks"]
            self.assertEqual(ws.max_row, 4)  # header + 3 tasks
