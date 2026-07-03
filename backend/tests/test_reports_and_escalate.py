"""Tests for the supplier-raised PO escalation + admin workload report.

DB-backed with an in-memory SQLite (production data untouched).
"""
from __future__ import annotations

import os
import unittest
from contextlib import contextmanager
from datetime import datetime, timedelta

os.environ.setdefault("DATABASE_URL", "sqlite:///./_test_reports.sqlite")

from sqlalchemy import create_engine, select  # noqa: E402
from sqlalchemy.orm import sessionmaker  # noqa: E402

from app.database import Base  # noqa: E402
from app.models import (  # noqa: E402,F401
    Asn,
    CommunicationMessage,
    CommunicationTask,
    Notification,
    ProcurementRecord,
    SupplierMaster,
    User,
)
from app.routers import portal as portal_router  # noqa: E402
from app.routers import reports as reports_router  # noqa: E402
from app.schemas.portal import PortalEscalateIn  # noqa: E402
from app.services import asn_service, user_service  # noqa: E402


@contextmanager
def _temp_db():
    engine = create_engine("sqlite:///:memory:", future=True)
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    db = Session()
    try:
        yield db
    finally:
        db.close()
        engine.dispose()


def _seed(db):
    sup = SupplierMaster(supplier_name="ACME TOOLS", is_active=True)
    db.add(sup)
    db.commit()
    db.refresh(sup)

    staff = user_service.create_user(
        db, email="admin@x.com", password="pw", full_name="Admin", role="admin"
    )
    emp = user_service.create_user(
        db, email="emp.e1@employee.local", password="pw", full_name="Desk Owner",
        role="employee", emp_code="E1", username="DESK1",
    )
    sup_user = user_service.create_user(
        db, email="orders@acme.com", password="pw", full_name="ACME",
        role="supplier", supplier_id=sup.id,
    )

    yesterday = datetime.utcnow() - timedelta(days=1)
    r1 = ProcurementRecord(
        crm_no="C1", supplier_po_no="PO-1", material_name="Drill",
        supplier_name="ACME TOOLS", owner_emp_code="E1",
        signal="RED", po_status="OPEN", shipment_date=yesterday, followup_count=2,
    )
    r2 = ProcurementRecord(
        crm_no="C2", supplier_po_no="PO-2", material_name="Bit",
        supplier_name="ACME TOOLS", owner_emp_code="E1",
        signal="GREEN", po_status="CLOSED",
        shipment_date=datetime.utcnow() + timedelta(days=5), followup_count=0,
    )
    db.add_all([r1, r2])
    db.commit()
    return sup, staff, emp, sup_user, r1, r2


class SupplierEscalateTests(unittest.TestCase):
    def test_escalate_flags_po_creates_task_message_and_is_idempotent(self):
        with _temp_db() as db:
            sup, staff, emp, sup_user, r1, _ = _seed(db)

            out = portal_router.escalate_po(
                "PO-1", PortalEscalateIn(reason="Line down — need this now"),
                user=sup_user, db=db,
            )
            self.assertTrue(out["ok"])
            self.assertFalse(out["already_escalated"])

            db.refresh(r1)
            self.assertEqual(r1.escalation_level, "ESCALATED")

            task = db.get(CommunicationTask, out["task_id"])
            self.assertEqual(task.task_source, "ESCALATION")
            self.assertEqual(task.assigned_by, "Supplier Portal")
            self.assertEqual(task.priority, "P0")
            # Routed to the PO's desk owner (emp_code E1).
            self.assertEqual(task.assigned_to_user_id, emp.id)

            msg = db.scalar(
                select(CommunicationMessage).where(
                    CommunicationMessage.mail_type == "PORTAL_ESCALATION"
                )
            )
            self.assertIsNotNone(msg)
            self.assertEqual(msg.supplier_po_no, "PO-1")

            notes = db.scalars(
                select(Notification).where(Notification.type == "SUPPLIER_ESCALATION")
            ).all()
            self.assertGreaterEqual(len(notes), 1)

            # Second click: no duplicate task.
            again = portal_router.escalate_po(
                "PO-1", PortalEscalateIn(reason=None), user=sup_user, db=db
            )
            self.assertTrue(again["already_escalated"])
            count = db.scalars(
                select(CommunicationTask).where(
                    CommunicationTask.task_source == "ESCALATION"
                )
            ).all()
            self.assertEqual(len(count), 1)

    def test_escalate_unknown_po_404s(self):
        from fastapi import HTTPException

        with _temp_db() as db:
            _, _, _, sup_user, _, _ = _seed(db)
            with self.assertRaises(HTTPException):
                portal_router.escalate_po(
                    "NOPE-99", PortalEscalateIn(), user=sup_user, db=db
                )


class WorkloadReportTests(unittest.TestCase):
    def test_per_user_per_supplier_and_overall_rollups(self):
        with _temp_db() as db:
            sup, staff, emp, sup_user, r1, _ = _seed(db)

            now = datetime.utcnow()
            db.add_all([
                CommunicationTask(
                    title="Chase PO-1", supplier_name="ACME TOOLS",
                    supplier_po_no="PO-1", status="TODO", priority="P1",
                    assigned_to_user_id=emp.id, assigned_to="Desk Owner",
                    due_date=now - timedelta(hours=3),
                ),
                CommunicationTask(
                    title="Done task", supplier_name="ACME TOOLS",
                    status="DONE", priority="P2",
                    assigned_to_user_id=emp.id, assigned_to="Desk Owner",
                ),
                CommunicationTask(title="Orphan", status="TODO", priority="P3"),
            ])
            asn_service.create_asn(
                db, supplier_id=sup.id, supplier_name="ACME TOOLS",
                supplier_po_no="PO-1", submit=True,
            )
            db.add(CommunicationMessage(
                direction="INCOMING", status="RECEIVED", channel="EMAIL",
                supplier_id=sup.id, supplier_name="ACME TOOLS",
                supplier_po_no="PO-1", subject="hi", body="b",
            ))
            db.commit()

            data = reports_router.workload_report(db=db)

            urow = next(u for u in data["users"] if u["user_id"] == emp.id)
            self.assertEqual(urow["pos"]["total"], 2)
            self.assertEqual(urow["pos"]["pending"], 1)   # PO-2 is CLOSED
            self.assertEqual(urow["pos"]["overdue"], 1)   # r1 shipped yesterday
            self.assertEqual(urow["pos"]["red"], 1)
            self.assertEqual(urow["tasks"]["open"], 1)
            self.assertEqual(urow["tasks"]["overdue"], 1)
            self.assertEqual(urow["tasks"]["done"], 1)

            # Staff admin exists in the list with empty workload.
            arow = next(u for u in data["users"] if u["user_id"] == staff.id)
            self.assertEqual(arow["pos"]["total"], 0)

            srow = next(
                s for s in data["suppliers"] if s["supplier_name"] == "ACME TOOLS"
            )
            self.assertEqual(srow["pos"]["total"], 2)
            self.assertEqual(srow["pos"]["pending"], 1)
            self.assertEqual(srow["worst_signal"], "RED")
            self.assertEqual(srow["tasks"]["open"], 1)
            self.assertEqual(srow["mails"]["incoming"], 1)
            self.assertEqual(srow["mails"]["unread"], 1)
            self.assertEqual(srow["asns"]["in_transit"], 1)

            self.assertEqual(data["overall"]["pos"]["total"], 2)
            self.assertEqual(data["overall"]["tasks"]["open"], 2)  # incl. orphan
            self.assertEqual(data["overall"]["unassigned_open_tasks"], 1)
            self.assertEqual(data["overall"]["unread_inbound"], 1)
            self.assertEqual(data["overall"]["asns_in_transit"], 1)
            self.assertEqual(data["overall"]["suppliers_active"], 1)

    def test_by_customer_groups_procurement_by_customer_name(self):
        with _temp_db() as db:
            _seed(db)
            # Two customers, one sharing a supplier PO number across suppliers.
            db.add_all([
                ProcurementRecord(
                    crm_no="CC1", supplier_po_no="000449", material_name="Drill A",
                    supplier_name="Vedant Tools Pvt Ltd", customer_name="ZANVAR GROUP",
                    signal="RED", po_status="OPEN",
                ),
                ProcurementRecord(
                    crm_no="CC2", supplier_po_no="000449", material_name="Bar B",
                    supplier_name="ALFA TOOLINGS", customer_name="ZANVAR GROUP",
                    signal="YELLOW", po_status="OPEN",
                ),
                ProcurementRecord(
                    crm_no="CC3", supplier_po_no="000501", material_name="Bit C",
                    supplier_name="Vedant Tools Pvt Ltd", customer_name="OTHER CUST",
                    signal="GREEN", po_status="OPEN",
                ),
            ])
            db.commit()

            data = reports_router.workload_report(db=db)
            customers = {c["customer_name"]: c for c in data["customers"]}
            self.assertIn("ZANVAR GROUP", customers)
            zg = customers["ZANVAR GROUP"]
            self.assertEqual(zg["pos"]["total"], 2)
            self.assertEqual(zg["suppliers"], 2)      # Vedant + Alfa on one customer
            self.assertEqual(zg["worst_signal"], "RED")
            self.assertEqual(data["overall"]["customers_active"], 2)


class WorkloadDetailTests(unittest.TestCase):
    def _seed_workload(self, db):
        sup, staff, emp, sup_user, r1, r2 = _seed(db)
        now = datetime.utcnow()
        db.add_all([
            CommunicationTask(
                title="Chase PO-1", supplier_name="ACME TOOLS",
                supplier_po_no="PO-1", status="TODO", priority="P1",
                assigned_to_user_id=emp.id, assigned_to="Desk Owner",
                due_date=now - timedelta(hours=3),
            ),
            CommunicationTask(
                title="Done task", supplier_name="ACME TOOLS",
                status="DONE", priority="P2",
                assigned_to_user_id=emp.id, assigned_to="Desk Owner",
                closed_at=now,
            ),
        ])
        asn_service.create_asn(
            db, supplier_id=sup.id, supplier_name="ACME TOOLS",
            supplier_po_no="PO-1", submit=True,
        )
        db.commit()
        return sup, emp

    def test_user_detail_shape(self):
        with _temp_db() as db:
            _, emp = self._seed_workload(db)
            d = reports_router.workload_user_detail(emp.id, db=db)
            self.assertEqual(d["user"]["emp_code"], "E1")
            self.assertEqual(d["pos"]["pending"], 1)
            self.assertEqual(len(d["pending_pos"]), 1)
            self.assertEqual(d["pending_pos"][0]["supplier_po_no"], "PO-1")
            self.assertGreaterEqual(d["pending_pos"][0]["days_overdue"], 1)
            self.assertEqual(len(d["open_tasks"]), 1)
            self.assertEqual(d["by_status"].get("TODO"), 1)
            self.assertEqual(d["by_priority"].get("P1"), 1)
            self.assertEqual(len(d["throughput"]), 14)
            # One task created + one completed today (last throughput bucket).
            self.assertGreaterEqual(d["throughput"][-1]["created"], 1)
            self.assertGreaterEqual(d["throughput"][-1]["completed"], 1)

    def test_supplier_detail_shape(self):
        with _temp_db() as db:
            sup, _ = self._seed_workload(db)
            d = reports_router.workload_supplier_detail(sup.id, db=db)
            self.assertEqual(d["supplier"]["supplier_name"], "ACME TOOLS")
            self.assertEqual(d["worst_signal"], "RED")
            self.assertEqual(d["pos"]["pending"], 1)
            self.assertEqual(len(d["asns"]), 1)
            self.assertEqual(d["tasks"]["open"], 1)

    def test_detail_404s(self):
        from fastapi import HTTPException

        with _temp_db() as db:
            _seed(db)
            with self.assertRaises(HTTPException):
                reports_router.workload_user_detail(99999, db=db)
            with self.assertRaises(HTTPException):
                reports_router.workload_supplier_detail(99999, db=db)

    def test_exports_are_valid_workbooks(self):
        import asyncio
        import io as _io

        from openpyxl import load_workbook

        with _temp_db() as db:
            sup, emp = self._seed_workload(db)

            def _bytes(resp):
                async def _collect():
                    return b"".join([chunk async for chunk in resp.body_iterator])
                return asyncio.run(_collect())

            full = load_workbook(_io.BytesIO(_bytes(reports_router.workload_export(db=db))))
            self.assertEqual(full.sheetnames, ["Overview", "Users", "Suppliers", "Customers"])
            users_ws = full["Users"]
            self.assertGreaterEqual(users_ws.max_row, 3)  # header + 2 users

            per_user = load_workbook(
                _io.BytesIO(_bytes(reports_router.workload_user_export(emp.id, db=db)))
            )
            self.assertEqual(per_user.sheetnames, ["Summary", "Pending POs", "Open Tasks"])
            self.assertEqual(per_user["Pending POs"].max_row, 2)  # header + PO-1

            per_sup = load_workbook(
                _io.BytesIO(_bytes(reports_router.workload_supplier_export(sup.id, db=db)))
            )
            self.assertEqual(
                per_sup.sheetnames, ["Summary", "Pending POs", "Open Tasks", "Shipments"]
            )
            self.assertEqual(per_sup["Shipments"].max_row, 2)  # header + 1 ASN


if __name__ == "__main__":
    unittest.main()
