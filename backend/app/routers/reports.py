"""Admin workload reports — per-user, per-supplier, and overall rollups,
plus per-entity detail drill-downs and Excel exports (meeting hand-outs).

One endpoint (`GET /api/reports/workload`) returns everything the admin
reporting dashboard needs in a single round-trip:

- ``users``      — every active internal account with the POs they own
                   (via ``ProcurementRecord.owner_emp_code == User.emp_code``)
                   and their task workload (``assigned_to_user_id``).
- ``suppliers``  — every active supplier with PO/signal/task/mail/ASN rollups
                   (procurement joins by UPPER(supplier_name); other tables
                   carry supplier_name too).
- ``overall``    — system-wide totals of the same measures.

Aggregations use SUM(CASE …) so they run identically on SQLite (tests) and
Postgres (prod). Mounted admin-only in main.py.
"""
from __future__ import annotations

import io
from datetime import datetime, timedelta
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import case, func, select
from sqlalchemy.orm import Session

from ..database import get_db
from ..models.asn import Asn
from ..models.communication_message import CommunicationMessage
from ..models.communication_task import CommunicationTask
from ..models.procurement import ProcurementRecord
from ..models.supplier import SupplierMaster
from ..models.user import User

router = APIRouter(prefix="/api/reports", tags=["reports"])

# Mirrors procurement_breakdown_service: a PO line is "pending" until its
# ERP status says it left the building.
_DELIVERED_STATUS = ("DISPATCHED", "DELIVERED", "CLOSED", "COMPLETED", "RECEIVED")
_ASN_CLOSED = ("DELIVERED", "CANCELLED")


def _one(x: Any) -> int:
    return int(x or 0)


def _po_measures() -> list[Any]:
    """Shared per-group PO aggregate columns (order matters — unpacked below)."""
    R = ProcurementRecord
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    pending = case(
        (R.po_status.is_(None), 1),
        (func.upper(R.po_status).notin_(_DELIVERED_STATUS), 1),
        else_=0,
    )
    overdue = case(
        (func.coalesce(R.shipment_date, today) < today, 1),
        else_=0,
    )
    sig = func.upper(func.coalesce(R.signal, ""))
    return [
        func.count(R.id),
        func.sum(pending),
        func.sum(overdue),
        func.sum(case((sig == "GREEN", 1), else_=0)),
        func.sum(case((sig == "YELLOW", 1), else_=0)),
        func.sum(case((sig == "RED", 1), else_=0)),
        func.sum(case((sig == "BLACK", 1), else_=0)),
        func.avg(func.coalesce(R.followup_count, 0)),
    ]


def _po_dict(row: Any) -> dict[str, Any]:
    return {
        "total": _one(row[0]),
        "pending": _one(row[1]),
        "overdue": _one(row[2]),
        "green": _one(row[3]),
        "yellow": _one(row[4]),
        "red": _one(row[5]),
        "black": _one(row[6]),
        "avg_followups": round(float(row[7] or 0), 1),
    }


def _task_measures() -> list[Any]:
    T = CommunicationTask
    now = datetime.utcnow()
    open_ = case((T.status != "DONE", 1), else_=0)
    overdue = case(
        (T.status != "DONE", case((func.coalesce(T.due_date, now) < now, 1), else_=0)),
        else_=0,
    )
    done = case((T.status == "DONE", 1), else_=0)
    due_today = case(
        (T.status != "DONE",
         case((T.due_date.between(now, now + timedelta(days=1)), 1), else_=0)),
        else_=0,
    )
    escalations = case(
        (T.task_source == "ESCALATION", case((T.status != "DONE", 1), else_=0)),
        else_=0,
    )
    return [
        func.count(T.id),
        func.sum(open_),
        func.sum(overdue),
        func.sum(done),
        func.sum(due_today),
        func.sum(escalations),
    ]


def _task_dict(row: Any) -> dict[str, Any]:
    return {
        "total": _one(row[0]),
        "open": _one(row[1]),
        "overdue": _one(row[2]),
        "done": _one(row[3]),
        "due_today": _one(row[4]),
        "escalations": _one(row[5]),
    }


_EMPTY_PO = {"total": 0, "pending": 0, "overdue": 0, "green": 0, "yellow": 0,
             "red": 0, "black": 0, "avg_followups": 0.0}
_EMPTY_TASK = {"total": 0, "open": 0, "overdue": 0, "done": 0, "due_today": 0,
               "escalations": 0}


@router.get("/workload")
def workload_report(db: Session = Depends(get_db)) -> dict[str, Any]:
    R, T, M, A = ProcurementRecord, CommunicationTask, CommunicationMessage, Asn

    # ── per-user ────────────────────────────────────────────────────────────
    users = list(
        db.scalars(
            select(User).where(User.is_active.is_(True), User.supplier_id.is_(None))
            .order_by(User.full_name.nullslast(), User.username)
        ).all()
    )
    po_by_emp = {
        (code or ""): row
        for code, *row in db.execute(
            select(R.owner_emp_code, *_po_measures())
            .where(R.owner_emp_code.is_not(None))
            .group_by(R.owner_emp_code)
        ).all()
    }
    task_by_user = {
        uid: row
        for uid, *row in db.execute(
            select(T.assigned_to_user_id, *_task_measures())
            .where(T.assigned_to_user_id.is_not(None))
            .group_by(T.assigned_to_user_id)
        ).all()
    }
    user_rows = []
    for u in users:
        po = po_by_emp.get((u.emp_code or "").strip())
        tk = task_by_user.get(u.id)
        user_rows.append({
            "user_id": u.id,
            "name": u.full_name or u.username or u.email,
            "role": u.role,
            "emp_code": u.emp_code,
            "last_login_at": u.last_login_at,
            "pos": _po_dict(po) if po else dict(_EMPTY_PO),
            "tasks": _task_dict(tk) if tk else dict(_EMPTY_TASK),
        })
    # Busiest first: open tasks, then pending POs.
    user_rows.sort(key=lambda r: (-r["tasks"]["open"], -r["pos"]["pending"], r["name"] or ""))

    # ── per-supplier (join key: UPPER(supplier_name)) ───────────────────────
    def _by_name(stmt) -> dict[str, Any]:
        return {(name or "").upper(): row for name, *row in db.execute(stmt).all()}

    po_by_sup = _by_name(
        select(R.supplier_name, *_po_measures())
        .where(R.supplier_name.is_not(None))
        .group_by(R.supplier_name)
    )
    task_by_sup = _by_name(
        select(T.supplier_name, *_task_measures())
        .where(T.supplier_name.is_not(None))
        .group_by(T.supplier_name)
    )
    mail_by_sup = _by_name(
        select(
            M.supplier_name,
            func.sum(case((M.direction == "INCOMING", 1), else_=0)),
            func.sum(case((M.direction == "OUTGOING", 1), else_=0)),
            func.sum(case(
                (M.direction == "INCOMING",
                 case((M.read_at.is_(None), 1), else_=0)),
                else_=0,
            )),
        )
        .where(M.supplier_name.is_not(None))
        .group_by(M.supplier_name)
    )
    asn_by_sup = _by_name(
        select(
            A.supplier_name,
            func.count(A.id),
            func.sum(case(
                (func.upper(A.status).notin_(_ASN_CLOSED + ("DRAFT",)), 1), else_=0,
            )),
            func.sum(case((func.upper(A.status) == "DELIVERED", 1), else_=0)),
        )
        .where(A.supplier_name.is_not(None))
        .group_by(A.supplier_name)
    )

    supplier_rows = []
    for s in db.scalars(
        select(SupplierMaster).where(SupplierMaster.is_active.is_(True))
        .order_by(SupplierMaster.supplier_name)
    ).all():
        key = (s.supplier_name or "").upper()
        po = po_by_sup.get(key)
        tk = task_by_sup.get(key)
        ml = mail_by_sup.get(key)
        an = asn_by_sup.get(key)
        po_d = _po_dict(po) if po else dict(_EMPTY_PO)
        worst = next(
            (c for c in ("BLACK", "RED", "YELLOW", "GREEN") if po_d[c.lower()] > 0),
            None,
        )
        supplier_rows.append({
            "supplier_id": s.id,
            "supplier_name": s.supplier_name,
            "worst_signal": worst,
            "pos": po_d,
            "tasks": _task_dict(tk) if tk else dict(_EMPTY_TASK),
            "mails": {
                "incoming": _one(ml[0]) if ml else 0,
                "outgoing": _one(ml[1]) if ml else 0,
                "unread": _one(ml[2]) if ml else 0,
            },
            "asns": {
                "total": _one(an[0]) if an else 0,
                "in_transit": _one(an[1]) if an else 0,
                "delivered": _one(an[2]) if an else 0,
            },
        })
    supplier_rows.sort(
        key=lambda r: (-r["pos"]["black"], -r["pos"]["red"], -r["pos"]["pending"],
                       r["supplier_name"] or "")
    )

    # ── per-customer (procurement only — the customer identity comes from the
    # PO feed; tasks/mail/ASN are not customer-scoped) ───────────────────────
    _has_customer = (R.customer_name.is_not(None), R.customer_name != "")
    cust_extra = {
        name: (nsup, nlines)
        for name, nsup, nlines in db.execute(
            select(
                R.customer_name,
                func.count(func.distinct(R.supplier_name)),
                func.count(R.id),
            ).where(*_has_customer).group_by(R.customer_name)
        ).all()
    }
    customer_rows = []
    for name, *po in db.execute(
        select(R.customer_name, *_po_measures())
        .where(*_has_customer)
        .group_by(R.customer_name)
    ).all():
        po_d = _po_dict(po)
        worst = next(
            (c for c in ("BLACK", "RED", "YELLOW", "GREEN") if po_d[c.lower()] > 0),
            None,
        )
        nsup, nlines = cust_extra.get(name, (0, 0))
        customer_rows.append({
            "customer_name": name,
            "worst_signal": worst,
            "pos": po_d,
            "suppliers": _one(nsup),
            "po_lines": _one(nlines),
        })
    customer_rows.sort(
        key=lambda r: (-r["pos"]["black"], -r["pos"]["red"], -r["pos"]["pending"],
                       r["customer_name"] or "")
    )

    # ── overall ─────────────────────────────────────────────────────────────
    po_all = db.execute(select(*_po_measures())).one()
    task_all = db.execute(select(*_task_measures())).one()
    overall = {
        "pos": _po_dict(po_all),
        "tasks": _task_dict(task_all),
        "suppliers_active": len(supplier_rows),
        "customers_active": len(customer_rows),
        "internal_users": len(user_rows),
        "unassigned_open_tasks": _one(db.scalar(
            select(func.count(T.id)).where(
                T.status != "DONE", T.assigned_to_user_id.is_(None)
            )
        )),
        "unread_inbound": _one(db.scalar(
            select(func.count(M.id)).where(
                M.direction == "INCOMING", M.read_at.is_(None)
            )
        )),
        "asns_in_transit": _one(db.scalar(
            select(func.count(A.id)).where(
                func.upper(A.status).notin_(_ASN_CLOSED + ("DRAFT",))
            )
        )),
        "generated_at": datetime.utcnow(),
    }
    return {
        "overall": overall,
        "users": user_rows,
        "suppliers": supplier_rows,
        "customers": customer_rows,
    }


# ── Detail drill-downs ───────────────────────────────────────────────────────
_ROW_CAP = 300


def _days_overdue(dt: datetime | None) -> int | None:
    if dt is None:
        return None
    delta = (datetime.utcnow() - dt).days
    return delta if delta > 0 else 0


def _pending_po_rows(db: Session, clause: Any) -> list[dict[str, Any]]:
    R = ProcurementRecord
    pending = (R.po_status.is_(None)) | (func.upper(R.po_status).notin_(_DELIVERED_STATUS))
    rows = db.scalars(
        select(R).where(clause, pending)
        .order_by(R.shipment_date.asc().nullslast())
        .limit(_ROW_CAP)
    ).all()
    return [
        {
            "procurement_record_id": r.id,
            "supplier_po_no": r.supplier_po_no,
            "supplier_name": r.supplier_name,
            "material_name": r.material_name,
            "qty": float(r.qty) if r.qty is not None else None,
            "uom": r.uom,
            "signal": r.signal,
            "po_status": r.po_status,
            "shipment_date": r.shipment_date,
            "days_overdue": _days_overdue(r.shipment_date),
            "followup_count": r.followup_count or 0,
            "commitment_date": r.commitment_date,
            "escalation_level": r.escalation_level,
        }
        for r in rows
    ]


def _open_task_rows(db: Session, clause: Any) -> list[dict[str, Any]]:
    T = CommunicationTask
    rows = db.scalars(
        select(T).where(clause, T.status != "DONE")
        .order_by(T.due_date.asc().nullslast(), T.priority.asc())
        .limit(_ROW_CAP)
    ).all()
    return [
        {
            "id": t.id,
            "title": t.title,
            "priority": t.priority,
            "status": t.status,
            "signal": t.signal,
            "task_source": t.task_source,
            "supplier_name": t.supplier_name,
            "supplier_po_no": t.supplier_po_no,
            "due_date": t.due_date,
            "days_overdue": _days_overdue(t.due_date),
            "progress_percent": t.progress_percent or 0,
        }
        for t in rows
    ]


def _task_breakdowns(db: Session, clause: Any) -> dict[str, Any]:
    """by_status / by_priority counts + avg cycle hours + 14-day throughput."""
    T = CommunicationTask
    by_status: dict[str, int] = {}
    for status_, n in db.execute(
        select(T.status, func.count(T.id)).where(clause).group_by(T.status)
    ).all():
        by_status[status_ or "?"] = int(n or 0)
    by_priority: dict[str, int] = {}
    for pri, n in db.execute(
        select(T.priority, func.count(T.id)).where(clause).group_by(T.priority)
    ).all():
        by_priority[pri or "?"] = int(n or 0)

    done = db.execute(
        select(T.created_at, T.closed_at).where(clause, T.closed_at.is_not(None))
        .order_by(T.closed_at.desc()).limit(500)
    ).all()
    cycles = [
        (c2 - c1).total_seconds() / 3600.0 for c1, c2 in done if c1 and c2 and c2 >= c1
    ]
    avg_cycle_hours = round(sum(cycles) / len(cycles), 1) if cycles else None

    # 14-day created/completed throughput (grouped in Python: cross-DB safe).
    since = datetime.utcnow() - timedelta(days=13)
    day0 = since.replace(hour=0, minute=0, second=0, microsecond=0)
    days = [(day0 + timedelta(days=i)).date() for i in range(14)]
    created_counts = dict.fromkeys(days, 0)
    completed_counts = dict.fromkeys(days, 0)
    for (created_at,) in db.execute(
        select(T.created_at).where(clause, T.created_at >= day0)
    ).all():
        if created_at and created_at.date() in created_counts:
            created_counts[created_at.date()] += 1
    for (closed_at,) in db.execute(
        select(T.closed_at).where(clause, T.closed_at >= day0)
    ).all():
        if closed_at and closed_at.date() in completed_counts:
            completed_counts[closed_at.date()] += 1
    throughput = [
        {"day": d.isoformat(), "created": created_counts[d], "completed": completed_counts[d]}
        for d in days
    ]
    return {
        "by_status": by_status,
        "by_priority": by_priority,
        "avg_cycle_hours": avg_cycle_hours,
        "throughput": throughput,
    }


def _user_detail(db: Session, user_id: int) -> dict[str, Any]:
    u = db.get(User, user_id)
    if u is None or u.supplier_id is not None:
        raise HTTPException(404, "User not found")
    R, T = ProcurementRecord, CommunicationTask
    emp_code = (u.emp_code or "").strip()
    po_clause = R.owner_emp_code == emp_code if emp_code else R.id.is_(None)
    task_clause = T.assigned_to_user_id == u.id

    po_row = db.execute(select(*_po_measures()).where(po_clause)).one()
    task_row = db.execute(select(*_task_measures()).where(task_clause)).one()
    return {
        "user": {
            "user_id": u.id,
            "name": u.full_name or u.username or u.email,
            "role": u.role,
            "emp_code": u.emp_code,
            "email": u.email,
            "last_login_at": u.last_login_at,
        },
        "pos": _po_dict(po_row),
        "tasks": _task_dict(task_row),
        **_task_breakdowns(db, task_clause),
        "pending_pos": _pending_po_rows(db, po_clause),
        "open_tasks": _open_task_rows(db, task_clause),
    }


def _supplier_detail(db: Session, supplier_id: int) -> dict[str, Any]:
    s = db.get(SupplierMaster, supplier_id)
    if s is None:
        raise HTTPException(404, "Supplier not found")
    R, T, M, A = ProcurementRecord, CommunicationTask, CommunicationMessage, Asn
    key = (s.supplier_name or "").upper()
    po_clause = func.upper(R.supplier_name) == key
    task_clause = func.upper(T.supplier_name) == key

    po_row = db.execute(select(*_po_measures()).where(po_clause)).one()
    task_row = db.execute(select(*_task_measures()).where(task_clause)).one()
    po_d = _po_dict(po_row)

    ml = db.execute(
        select(
            func.sum(case((M.direction == "INCOMING", 1), else_=0)),
            func.sum(case((M.direction == "OUTGOING", 1), else_=0)),
            func.sum(case(
                (M.direction == "INCOMING", case((M.read_at.is_(None), 1), else_=0)),
                else_=0,
            )),
        ).where(func.upper(M.supplier_name) == key)
    ).one()
    incoming, outgoing = _one(ml[0]), _one(ml[1])
    asns = db.scalars(
        select(A).where(func.upper(A.supplier_name) == key)
        .order_by(A.created_at.desc()).limit(_ROW_CAP)
    ).all()
    return {
        "supplier": {
            "supplier_id": s.id,
            "supplier_name": s.supplier_name,
            "is_active": s.is_active,
        },
        "worst_signal": next(
            (c for c in ("BLACK", "RED", "YELLOW", "GREEN") if po_d[c.lower()] > 0), None
        ),
        "pos": po_d,
        "tasks": _task_dict(task_row),
        **_task_breakdowns(db, task_clause),
        "mails": {
            "incoming": incoming,
            "outgoing": outgoing,
            "unread": _one(ml[2]),
            "response_rate": round(incoming / outgoing, 2) if outgoing else None,
        },
        "asns": [
            {
                "id": a.id,
                "asn_no": a.asn_no,
                "supplier_po_no": a.supplier_po_no,
                "status": a.status,
                "status_label": a.status_label,
                "progress_percent": a.progress_percent,
                "carrier_name": a.carrier_name,
                "tracking_no": a.tracking_no,
                "eta": a.eta,
                "alert": a.alert,
            }
            for a in asns
        ],
        "pending_pos": _pending_po_rows(db, po_clause),
        "open_tasks": _open_task_rows(db, task_clause),
    }


@router.get("/workload/users/{user_id}")
def workload_user_detail(user_id: int, db: Session = Depends(get_db)) -> dict[str, Any]:
    return _user_detail(db, user_id)


@router.get("/workload/suppliers/{supplier_id}")
def workload_supplier_detail(supplier_id: int, db: Session = Depends(get_db)) -> dict[str, Any]:
    return _supplier_detail(db, supplier_id)


# ── Excel exports ────────────────────────────────────────────────────────────
def _fmt_cell(v: Any) -> Any:
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M")
    return v


def _add_sheet(wb: Any, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    from openpyxl.styles import Font

    ws = wb.create_sheet(title=title[:31])
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in rows:
        ws.append([_fmt_cell(v) for v in row])
    ws.freeze_panes = "A2"
    for idx, h in enumerate(headers, start=1):
        width = max(len(str(h)) + 2, 12)
        for row in rows[:50]:
            if idx - 1 < len(row) and row[idx - 1] is not None:
                width = max(width, min(len(str(row[idx - 1])) + 2, 48))
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width


def _workbook_bytes(wb: Any) -> bytes:
    # Drop openpyxl's default empty sheet.
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_response(data: bytes, filename: str) -> StreamingResponse:
    return StreamingResponse(
        iter([data]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


_PENDING_PO_HEADERS = [
    "PO No", "Supplier", "Material", "Qty", "UOM", "Signal", "PO Status",
    "Shipment Date", "Days Overdue", "Follow-ups", "Commitment Date", "Escalation",
]
_OPEN_TASK_HEADERS = [
    "Task", "Priority", "Status", "Signal", "Source", "Supplier", "PO No",
    "Due Date", "Days Overdue", "Progress %",
]


def _pending_po_sheet_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    return [
        [r["supplier_po_no"], r["supplier_name"], r["material_name"], r["qty"], r["uom"],
         r["signal"], r["po_status"], r["shipment_date"], r["days_overdue"],
         r["followup_count"], r["commitment_date"], r["escalation_level"]]
        for r in rows
    ]


def _open_task_sheet_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    return [
        [t["title"], t["priority"], t["status"], t["signal"], t["task_source"],
         t["supplier_name"], t["supplier_po_no"], t["due_date"], t["days_overdue"],
         t["progress_percent"]]
        for t in rows
    ]


@router.get("/workload/export")
def workload_export(db: Session = Depends(get_db)) -> StreamingResponse:
    from openpyxl import Workbook

    data = workload_report(db=db)
    wb = Workbook()

    o = data["overall"]
    _add_sheet(wb, "Overview", ["Metric", "Value"], [
        ["Generated", o["generated_at"]],
        ["Internal users", o["internal_users"]],
        ["Active suppliers", o["suppliers_active"]],
        ["Active customers", o.get("customers_active", 0)],
        ["PO lines", o["pos"]["total"]],
        ["Pending POs", o["pos"]["pending"]],
        ["Overdue POs", o["pos"]["overdue"]],
        ["Green / Yellow / Red / Black",
         f'{o["pos"]["green"]} / {o["pos"]["yellow"]} / {o["pos"]["red"]} / {o["pos"]["black"]}'],
        ["Open tasks", o["tasks"]["open"]],
        ["Overdue tasks", o["tasks"]["overdue"]],
        ["Tasks done", o["tasks"]["done"]],
        ["Open escalations", o["tasks"]["escalations"]],
        ["Unassigned open tasks", o["unassigned_open_tasks"]],
        ["Unread inbound mail", o["unread_inbound"]],
        ["ASNs in transit", o["asns_in_transit"]],
    ])
    _add_sheet(
        wb, "Users",
        ["User", "Role", "Emp Code", "PO Lines", "Pending POs", "Overdue POs",
         "Red POs", "Black POs", "Open Tasks", "Overdue Tasks", "Due Today",
         "Done", "Open Escalations", "Last Login"],
        [
            [u["name"], u["role"], u["emp_code"], u["pos"]["total"], u["pos"]["pending"],
             u["pos"]["overdue"], u["pos"]["red"], u["pos"]["black"], u["tasks"]["open"],
             u["tasks"]["overdue"], u["tasks"]["due_today"], u["tasks"]["done"],
             u["tasks"]["escalations"], u["last_login_at"]]
            for u in data["users"]
        ],
    )
    _add_sheet(
        wb, "Suppliers",
        ["Supplier", "Worst Signal", "PO Lines", "Pending POs", "Overdue POs",
         "Red POs", "Black POs", "Avg Follow-ups", "Open Tasks", "Open Escalations",
         "Mail In", "Mail Out", "Unread", "ASNs", "In Transit", "Delivered"],
        [
            [s["supplier_name"], s["worst_signal"], s["pos"]["total"], s["pos"]["pending"],
             s["pos"]["overdue"], s["pos"]["red"], s["pos"]["black"],
             s["pos"]["avg_followups"], s["tasks"]["open"], s["tasks"]["escalations"],
             s["mails"]["incoming"], s["mails"]["outgoing"], s["mails"]["unread"],
             s["asns"]["total"], s["asns"]["in_transit"], s["asns"]["delivered"]]
            for s in data["suppliers"]
        ],
    )
    _add_sheet(
        wb, "Customers",
        ["Customer", "Worst Signal", "Suppliers", "PO Lines", "Pending POs",
         "Overdue POs", "Green", "Yellow", "Red", "Black"],
        [
            [c["customer_name"], c["worst_signal"], c["suppliers"], c["pos"]["total"],
             c["pos"]["pending"], c["pos"]["overdue"], c["pos"]["green"],
             c["pos"]["yellow"], c["pos"]["red"], c["pos"]["black"]]
            for c in data.get("customers", [])
        ],
    )
    stamp = datetime.utcnow().strftime("%Y-%m-%d")
    return _xlsx_response(_workbook_bytes(wb), f"workload-report-{stamp}.xlsx")


def _entity_workbook(detail: dict[str, Any], summary_rows: list[list[Any]]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    _add_sheet(wb, "Summary", ["Metric", "Value"], summary_rows)
    _add_sheet(wb, "Pending POs", _PENDING_PO_HEADERS,
               _pending_po_sheet_rows(detail["pending_pos"]))
    _add_sheet(wb, "Open Tasks", _OPEN_TASK_HEADERS,
               _open_task_sheet_rows(detail["open_tasks"]))
    if "asns" in detail:
        _add_sheet(
            wb, "Shipments",
            ["ASN", "PO No", "Status", "Progress %", "Carrier", "Tracking", "ETA", "Alert"],
            [
                [a["asn_no"], a["supplier_po_no"], a["status_label"] or a["status"],
                 a["progress_percent"], a["carrier_name"], a["tracking_no"],
                 a["eta"], "YES" if a["alert"] else ""]
                for a in detail["asns"]
            ],
        )
    return _workbook_bytes(wb)


@router.get("/workload/users/{user_id}/export")
def workload_user_export(user_id: int, db: Session = Depends(get_db)) -> StreamingResponse:
    d = _user_detail(db, user_id)
    u, p, t = d["user"], d["pos"], d["tasks"]
    rows = [
        ["User", u["name"]], ["Role", u["role"]], ["Emp code", u["emp_code"]],
        ["Generated", datetime.utcnow()],
        ["PO lines owned", p["total"]], ["Pending POs", p["pending"]],
        ["Overdue POs", p["overdue"]],
        ["Green / Yellow / Red / Black",
         f'{p["green"]} / {p["yellow"]} / {p["red"]} / {p["black"]}'],
        ["Open tasks", t["open"]], ["Overdue tasks", t["overdue"]],
        ["Due today", t["due_today"]], ["Done", t["done"]],
        ["Open escalations", t["escalations"]],
        ["Avg task cycle (hours)", d["avg_cycle_hours"]],
    ]
    slug = "".join(ch for ch in (u["name"] or "user") if ch.isalnum() or ch in " -_").strip().replace(" ", "-")[:40]
    return _xlsx_response(_entity_workbook(d, rows), f"workload-{slug or 'user'}.xlsx")


@router.get("/workload/suppliers/{supplier_id}/export")
def workload_supplier_export(supplier_id: int, db: Session = Depends(get_db)) -> StreamingResponse:
    d = _supplier_detail(db, supplier_id)
    s, p, t, m = d["supplier"], d["pos"], d["tasks"], d["mails"]
    rows = [
        ["Supplier", s["supplier_name"]], ["Generated", datetime.utcnow()],
        ["Worst signal", d["worst_signal"]],
        ["PO lines", p["total"]], ["Pending POs", p["pending"]],
        ["Overdue POs", p["overdue"]],
        ["Green / Yellow / Red / Black",
         f'{p["green"]} / {p["yellow"]} / {p["red"]} / {p["black"]}'],
        ["Avg follow-ups per line", p["avg_followups"]],
        ["Open tasks", t["open"]], ["Open escalations", t["escalations"]],
        ["Mail incoming / outgoing", f'{m["incoming"]} / {m["outgoing"]}'],
        ["Unread inbound", m["unread"]],
        ["Reply rate (in/out)", m["response_rate"]],
        ["Shipments (total)", len(d["asns"])],
    ]
    slug = "".join(ch for ch in (s["supplier_name"] or "supplier") if ch.isalnum() or ch in " -_").strip().replace(" ", "-")[:40]
    return _xlsx_response(_entity_workbook(d, rows), f"workload-{slug or 'supplier'}.xlsx")
