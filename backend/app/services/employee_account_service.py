"""Internal employee portal login provisioning.

Employees have **no email** — they log in with their CRM login id (e.g. "PRAMOD")
and are scoped to POs where ``owner_emp_code == User.emp_code``. Accounts are
provisioned from the Hariom employee sheet: username = ``EMPLOYEE_LOGIN_ID``,
emp_code = ``EMPLOYEE_ID``, a random temp password, and ``must_change_password``
so the first login forces a change. Credentials are handed back to the admin as
an Excel (there is no email channel). Pure service layer: no FastAPI imports.
"""
from __future__ import annotations

import io
import logging
import re
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from ..core.roles import Role
from ..models.user import User
from . import user_service
from .supplier_account_service import generate_temp_password
from .user_service import EmailTakenError, UsernameTakenError

log = logging.getLogger(__name__)

# Synthetic email domain — employees never receive mail here; it only satisfies
# the NOT NULL/UNIQUE `users.email` column. The username is the real login key.
_PLACEHOLDER_DOMAIN = "employee.local"


def _placeholder_email(username: str, emp_code: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "", (username or "").lower()) or "emp"
    return f"{slug}.{emp_code}@{_PLACEHOLDER_DOMAIN}"


def _full_name(row: dict[str, Any]) -> str | None:
    parts = [
        str(row.get(k) or "").strip()
        for k in ("EMPLOYEE_FIRST_NAME", "EMPLOYEE_MIDDLE_NAME", "EMPLOYEE_LAST_NAME")
    ]
    name = " ".join(p for p in parts if p)
    return name or None


def list_employee_logins(db: Session) -> list[User]:
    return list(
        db.scalars(
            select(User).where(User.emp_code.is_not(None)).order_by(User.username)
        ).all()
    )


def parse_employee_sheet(content: bytes) -> list[dict[str, Any]]:
    """Read the first worksheet into a list of header→value dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[dict[str, Any]] = []
    for r in rows[1:]:
        if not any(v not in (None, "") for v in r):
            continue
        out.append(dict(zip(header, r)))
    return out


def provision_from_rows(db: Session, rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Create an employee login per sheet row. Idempotent: existing usernames are
    kept (reactivated if disabled), never re-passworded. Returns a summary with
    temp passwords for the new accounts (the admin distributes them)."""
    created: list[dict[str, Any]] = []
    reactivated: list[str] = []
    conflicts: list[dict[str, str]] = []
    skipped: list[dict[str, str]] = []

    for row in rows:
        emp_id = row.get("EMPLOYEE_ID")
        login_id = str(row.get("EMPLOYEE_LOGIN_ID") or "").strip()
        if not emp_id or not login_id:
            skipped.append({"row": str(row.get("EMPLOYEE_ID") or row), "reason": "missing id/login"})
            continue
        emp_code = str(emp_id).strip()
        full_name = _full_name(row)

        existing = user_service.get_by_username(db, login_id)
        if existing is not None:
            if existing.emp_code == emp_code:
                if not existing.is_active:
                    existing.is_active = True
                    reactivated.append(login_id)
            else:
                conflicts.append({"username": login_id, "reason": "username already in use"})
            continue

        temp_password = generate_temp_password()
        try:
            user_service.create_user(
                db,
                email=_placeholder_email(login_id, emp_code),
                password=temp_password,
                full_name=full_name,
                role=Role.EMPLOYEE,
                is_active=True,
                emp_code=emp_code,
                username=login_id,
                must_change_password=True,
                commit=False,
            )
        except (EmailTakenError, UsernameTakenError) as exc:
            conflicts.append({"username": login_id, "reason": str(exc)})
            continue
        created.append(
            {"username": login_id, "full_name": full_name, "emp_code": emp_code, "temp_password": temp_password}
        )

    db.commit()
    log.info(
        "[employee] provision: created=%d reactivated=%d conflicts=%d skipped=%d",
        len(created), len(reactivated), len(conflicts), len(skipped),
    )
    return {
        "created": created,
        "reactivated": reactivated,
        "conflicts": conflicts,
        "skipped": skipped,
    }


def create_employee(
    db: Session, *, username: str, full_name: str | None, emp_code: str | None
) -> dict[str, Any]:
    """Admin 'add user': create a single employee login with a temp password."""
    username = username.strip()
    emp_code = (emp_code or "").strip() or None
    user = user_service.create_user(
        db,
        email=_placeholder_email(username, emp_code or username),
        password=(temp_password := generate_temp_password()),
        full_name=full_name,
        role=Role.EMPLOYEE,
        is_active=True,
        emp_code=emp_code,
        username=username,
        must_change_password=True,
        commit=True,
    )
    return {
        "id": user.id,
        "username": username,
        "full_name": full_name,
        "emp_code": emp_code,
        "temp_password": temp_password,
    }


def reset_employee_password(db: Session, user: User) -> dict[str, Any]:
    """Admin reset: new temp password + force change on next login."""
    temp_password = generate_temp_password()
    user_service.set_password(db, user, temp_password, must_change=True)
    return {
        "id": user.id,
        "username": user.username,
        "full_name": user.full_name,
        "temp_password": temp_password,
    }


def credentials_workbook(items: list[dict[str, Any]]) -> bytes:
    """Build a 'Name / Username / Temporary Password' .xlsx for admin hand-out."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Credentials"
    ws.append(["Name", "Username", "Temporary Password"])
    for it in items:
        ws.append([it.get("full_name") or "", it.get("username") or "", it.get("temp_password") or ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
